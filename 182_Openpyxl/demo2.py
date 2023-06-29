# creating a workbook
from openpyxl import Workbook
wb = Workbook()
ws = wb.active # This sheet is active

# create new worksheet
ws1 = wb.create_sheet("mysheet1") # insert at end
ws2 = wb.create_sheet("mysheet2",0) # insert at first position
ws3 = wb.create_sheet("mysheet3",-1) # insert at penultimate position

ws.title = "New Title" # Active sheet's title renamed

print(wb.sheetnames)

c = ws['A4']
ws['A4'] = 4
d = ws.cell(row=4, column=2, value=10)
print(d)


wb.save("test.xlsx")